#!/usr/bin/env python3
#-*- coding: utf-8 -*-

# http://yuizho.hatenablog.com/entry/2013/04/16/234410
# http://mochio326.blog.shinobi.jp/python/python%E3%81%8B%E3%82%89excel%E3%82%92%E6%93%8D%E4%BD%9C%E3%81%99%E3%82%8B

import sys
import glob
import xlrd
import openpyxl
import argparse
import tqdm
import csv
import warnings
import re

class xl_grepper(object):
	def __init__(self, args):
		self._writer = csv.writer(open("result.csv", 'w', encoding=args.encoding), lineterminator=args.lineterminator)
		self._sp = re.compile(args.sheetfilter)
		self._rfmin = int(args.rowfilter.split("-")[0])
		self._rfmax = int(args.rowfilter.split("-")[1])
		self._cfmin = int(args.colfilter.split("-")[0])
		self._cfmax = int(args.colfilter.split("-")[1])
	
	# 数字をアルファベットに変換 1 -> A 27 -> AA
	def num2char(num):
		quotient, remainder = divmod(num, 26)
		chars = ''
		if quotient > 0: chars = chr(quotient + 64)
		if remainder > 0: chars = chars + chr(remainder + 64)
		return chars
	
	# アルファベットを数字に変換 A -> 1  AA -> 27
	def char2num(chars):
		num = 0
		for c in chars: num = num * 26 + (ord(c) - 64)
		return num
	
	def _sheetFilter(self, sheetnm):
		return self._sp.match(sheetnm)
		
	def _read(self, bookpath):
		pass
	
	def _judge(self, value):
		if value.find(self.keyword) != -1:
			return True
		return False
	
	def _action(self, bookpath, sheetnm, cellpos):
		self._writer.writerow([bookpath, sheetnm, cellpos])
	
	def do(self, keyword, bookpath):
		self.keyword = keyword
		self._read(file)

# For Excel2000/2002/2003
class xls_grepper(xl_grepper):
	def __init__(self, args):
		super().__init__(args)
		self.ext = '.xls'

	def _read(self, bookpath):
		book = xlrd.open_workbook(bookpath)
		for sheet in book.sheets():
			for col in range(sheet.ncols):
				for row in range(sheet.nrows):
					if self._judge(sheet.cell(row, col).value):
						self._action(bookpath, sheet.name, str(row) + str(row))

# For Excel2007/2010/2013
class xlsx_grepper(xl_grepper):
	def __init__(self, args):
		super().__init__(args)
		self.ext = '.xls[x|m]'

	def _read(self, bookpath):
		with warnings.catch_warnings():
			warnings.simplefilter("ignore")
			book = openpyxl.load_workbook(bookpath, data_only=True)
			for sheet in book:
				if not self._sheetFilter(sheet.title) : continue
				_rowmax = self._rfmax
				_colmax = self._cfmax
				if sheet.max_row < self._rfmax: _rowmax = sheet.max_row
				if sheet.max_column < self._cfmax: _colmax = sheet.max_column
				for rowidx in range(self._rfmin, _rowmax) :
					for colidx in range(self._cfmin, _colmax):
						if sheet.cell(row=rowidx, column=colidx).value != None:
							if self._judge(str(sheet.cell(row=rowidx, column=colidx).value)):
								self._action(bookpath, sheet.title, str(colidx) + str(rowidx))

if __name__ == '__main__':
	
	# Argument setting
	parser = argparse.ArgumentParser()
	parser.add_argument('keyword', help='grep keyword')
	parser.add_argument('-f', '--filename', default='*', help='grep target file name. Regular expression possible. Not include the extension.')
	parser.add_argument('-o', '--old', action='store_true', help='grep old version Excel Files. (ex. ".xls")' )
	parser.add_argument('-b', '--basedir', default='.', help='grep target base directory. Regular expression impossible.')
	parser.add_argument('-e', '--encoding', default='shift-jis', help='Client charactor encoding.')
	parser.add_argument('-l', '--lineterminator', default='\r\n', help='Client line terminator.')
	parser.add_argument('-sf', '--sheetfilter', default='.*', help='Filtering sheet name. Regular expression possible.')
	parser.add_argument('-rf', '--rowfilter', default='1-99999', help='Filtering row range.(ex. 1-30)')
	parser.add_argument('-cf', '--colfilter', default='1-99999', help='Filtering col range.(ex. 1-15)')
	
	# Check arguments
	args = parser.parse_args()
	print(args)
	
	# Create an instance of the "grepper" by argument
	if args.old:
		grepper = xls_grepper(args)
	else:
		grepper = xlsx_grepper(args)
	
	# Grep file specified by argument
	filter = args.basedir + '/**/[!~]' + args.filename + grepper.ext
	print(filter)
	files = glob.glob(filter, recursive=True)
	for file in tqdm.tqdm(files):
		try:
			grepper.do(args.keyword, file)
		except:
			print(file)
			raise
